import openpyxl
from ColorArrayArtist import ColorArrayArtist


class ExcelArtist(ColorArrayArtist):
    """
    エクセルでお絵かきするためのクラス
    """

    def __init__(self, filename, cell_size=10):
        super().__init__()
        self.cell_size = cell_size
        self.filename = filename
        self.x = 0
        self.y = 0

    def open(self, color_array):
        self.file = openpyxl.Workbook()
        sheet = self.file.active
        for y in range(len(color_array)):
            # pylint: disable=maybe-no-member
            sheet.row_dimensions[y + 1].height = self.cell_size / 2
            for x in range(len(color_array[0])):
                letter = openpyxl.utils.get_column_letter(x + 1)
                sheet.column_dimensions[letter].width = self.cell_size / 11.07
            # pylint: enable=maybe-no-member

    def put_pixel(self, rgb):
        key = openpyxl.utils.get_column_letter(self.x + 1) + str(self.y + 1)
        self.file.active[key].fill = openpyxl.styles.PatternFill(
            patternType='solid', fgColor=self.x11_from_rgb(rgb))
        self.x += 1

    def new_line(self):
        self.x = 0
        self.y += 1

    def close(self):
        self.file.save(filename=self.filename)

    @staticmethod
    def x11_from_rgb(rgb):
        """
        色のRGB表記(R, G, B)をX11表記(#ARGB)に直す
        r,g,bは0から255
        aは255で固定
        """
        r_str = hex(rgb[0])[2:].rjust(2, '0')
        g_str = hex(rgb[1])[2:].rjust(2, '0')
        b_str = hex(rgb[2])[2:].rjust(2, '0')
        return "ff" + r_str + g_str + b_str
