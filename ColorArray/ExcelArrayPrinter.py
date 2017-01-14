from typing import List

import openpyxl

from ColorArray.ColorArrayPrinter import ColorArrayPrinter
from Painter.Figure import ColorArray, Point


class ExcelArrayPrinter(ColorArrayPrinter):
    """ エクセルでお絵かきするためのクラス """

    def __init__(self, filename: str, cell_size: int = 10) -> None:
        super().__init__()
        self.cell_size = cell_size
        self.filename = filename + ".xlsx"
        self.file = None

    def open(self, color_array: ColorArray) -> None:
        """ 初期設定を行う """
        self.file = openpyxl.Workbook()
        sheet = self.file.active
        for y in range(len(color_array)):
            sheet.row_dimensions[y + 1].height = self.cell_size / 2
            for x in range(len(color_array[0])):
                letter = openpyxl.utils.get_column_letter(x + 1)
                sheet.column_dimensions[letter].width = self.cell_size / 11.07

    def put_pixel(self, point: Point) -> None:
        """ pointを描画する """
        key = openpyxl.utils.get_column_letter(point.x + 1) + str(point.y + 1)
        self.file.active[key].fill = openpyxl.styles.PatternFill(
                patternType='solid', fgColor=self.x11_from_rgb(point.rgb))

    def new_line(self) -> None:
        """ 行を改める """
        pass

    def close(self) -> None:
        """ ファイルを閉じる """
        self.file.save(filename=self.filename)

    @staticmethod
    def x11_from_rgb(rgb: List[int]) -> str:
        """
        色のRGB表記(R, G, B)をX11表記(#ARGB)に直す
        aはffで固定
        """
        r_str = hex(rgb[0])[2:].rjust(2, '0')
        g_str = hex(rgb[1])[2:].rjust(2, '0')
        b_str = hex(rgb[2])[2:].rjust(2, '0')
        return "ff" + r_str + g_str + b_str
